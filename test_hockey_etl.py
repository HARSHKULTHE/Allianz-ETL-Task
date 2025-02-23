from openpyxl import Workbook
import pytest
from bs4 import BeautifulSoup
from JsontoExcel import write_nhl_stats_sheet, compute_winners_and_losers, extract_team_stats  


SAMPLE_HTML = """
<table>
    <tr class="team">
        <td>1995</td>
        <td>Detroit Red Wings</td>
        <td>57</td>
        <td>13</td>
    </tr>
    <tr class="team">
        <td>1995</td>
        <td>Los Angeles Kings</td>
        <td>20</td>
        <td>52</td>
    </tr>
</table>
"""

def test_write_nhl_stats_sheets():
    soup = BeautifulSoup(SAMPLE_HTML, "html.parser")
    team_stats = extract_team_stats(str(soup)) 

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "NHL_Stats"

    write_nhl_stats_sheet(worksheet, team_stats) 

    assert worksheet.cell(row=1, column=1).value == "Year"
    assert worksheet.cell(row=1, column=2).value == "Team Name"
    assert worksheet.cell(row=1, column=3).value == "Wins"
    assert worksheet.cell(row=1, column=4).value == "Losses"

    assert worksheet.cell(row=2, column=1).value == "1995"
    assert worksheet.cell(row=2, column=2).value == "Detroit Red Wings"
    assert worksheet.cell(row=2, column=3).value == 57
    assert worksheet.cell(row=2, column=4).value == 13


def test_compute_winners_and_losers():
    """Test winner and loser calculation logic."""
    team_data = [
        {"year": "1995", "team_name": "Detroit Red Wings", "wins": 57, "losses": 13},
        {"year": "1995", "team_name": "Los Angeles Kings", "wins": 20, "losses": 52}
    ]

    yearly_summary = compute_winners_and_losers(team_data)

    assert yearly_summary["1995"]["winner"] == ("Detroit Red Wings", 57)
    assert yearly_summary["1995"]["loser"] == ("Los Angeles Kings", 20)
